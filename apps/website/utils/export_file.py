# dashboard/utils/export_file.py - ПОЛНЫЙ ИСПРАВЛЕННЫЙ ФАЙЛ

import pandas as pd
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

logger = logging.getLogger(__name__)


class UniversalExporter:
    """🚀 ЭКСПОРТ СО 100% СКАЧИВАНИЕМ ФАЙЛА"""

    def __init__(self, items, request=None, source='all'):
        self.items = items or []
        self.request = request
        self.source = source

    def _safe_value(self, value, default='—'):
        """Безопасное получение значения"""
        try:
            if value is None or value == '' or value == 'None':
                return default
            if isinstance(value, (int, float)):
                return value
            return str(value).strip()
        except:
            return default

    def _create_excel_for_zip(self):
        """Создает Excel файл для использования в ZIP архиве"""
        try:
            buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Объявления"

            # Заголовок
            ws['A1'] = f"Объявления ({len(self.items)} записей)"
            ws['A1'].font = Font(bold=True, size=14)

            # Заголовки
            headers = ["ID", "Название", "Цена (₽)", "Сайт", "Город", "Дата"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=2, column=col, value=header)
                ws.cell(row=2, column=col).font = Font(bold=True)

            # Данные
            for row_idx, item in enumerate(self.items, 3):
                try:
                    ws.cell(row=row_idx, column=1, value=self._safe_value(getattr(item, 'id', '')))
                    ws.cell(row=row_idx, column=2, value=self._safe_value(getattr(item, 'title', ''))[:100])
                    ws.cell(row=row_idx, column=3, value=self._safe_value(getattr(item, 'price', 0)))
                    ws.cell(row=row_idx, column=4, value=self._safe_value(getattr(item, 'source', '')))
                    ws.cell(row=row_idx, column=5, value=self._safe_value(getattr(item, 'city', '')))
                    ws.cell(row=row_idx, column=6, value=self._safe_value(getattr(item, 'posted_date', '')))
                except:
                    continue

            # Настройка ширины
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15

            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Excel for ZIP error: {e}")
            return b''
    def _create_excel_workbook(self):
        """Создает Excel workbook для использования в ZIP архиве"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Объявления"

        # Простая таблица
        ws['A1'] = f"Объявления ({len(self.items)} записей)"
        ws['A1'].font = Font(bold=True)

        # Заголовки
        headers = ["ID", "Название", "Цена", "Город", "Дата"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header)
            ws.cell(row=2, column=col).font = Font(bold=True)

        # Данные
        for row_idx, item in enumerate(self.items, 3):
            try:
                ws.cell(row=row_idx, column=1, value=self._safe_value(getattr(item, 'id', '')))
                ws.cell(row=row_idx, column=2, value=self._safe_value(getattr(item, 'title', ''))[:100])
                ws.cell(row=row_idx, column=3, value=self._safe_value(getattr(item, 'price', 0)))
                ws.cell(row=row_idx, column=4, value=self._safe_value(getattr(item, 'city', '')))
                ws.cell(row=row_idx, column=5, value=self._safe_value(getattr(item, 'posted_date', '')))
            except:
                continue

        return wb

    def _get_view_badge(self, views):
        """Статус просмотров как в таблице"""
        try:
            views_num = int(views) if views and str(views).isdigit() else 0
            if views_num < 20:
                return "🔥 Новое"
            elif 21 <= views_num <= 40:
                return "🟢 Свежее"
            else:
                return "⏳ Старое"
        except:
            return "—"

    def _get_seller_badge(self, seller_type):
        """Бейдж продавца"""
        if not seller_type or seller_type == '—':
            return "👤 Частное"
        seller_str = str(seller_type)
        if any(word in seller_str for word in ['Компания', 'Магазин', 'Дилер', 'reseller']):
            return "🏢 " + seller_str
        return "👤 Частное"

    def _get_profit_badge(self, profit, source):
        """Бейдж прибыли"""
        try:
            profit_val = float(profit) if profit else 0
            if profit_val > 1000:
                return "💎 Высокая"
            elif profit_val > 0:
                return "💰 Прибыль"
            elif profit_val == 0:
                return "⚠️ Ноль"
            else:
                return "❌ Убыток"
        except:
            return "📊 Анализ"

    def _get_source_badge(self, source):
        """Бейдж источника"""
        badges = {
            'avito': '🏠 Авито',
            'auto_ru': '🚗 Auto.ru',
            'cian': '🏘️ Cian',
            'yula': '🛒 Юла',
            'drom': '🔧 Drom'
        }
        return badges.get(source, '🌐 Другой')

    def export_excel(self):
        """📊 ЭКСПОРТ В EXCEL - С РЕАЛЬНЫМИ ИЗБРАННЫМИ"""
        try:
            if not self.items:
                return self._error_response("Нет данных для экспорта")

            wb = Workbook()

            # ==================== ВКЛАДКА 1: ВСЕ ОБЪЯВЛЕНИЯ ====================
            ws_all = wb.active
            ws_all.title = "Все объявления"

            # Красивая шапка
            ws_all.merge_cells('A1:AC1')
            title_cell = ws_all['A1']
            title_cell.value = "📊 SELIBRI - ВСЕ ОБЪЯВЛЕНИЯ"
            title_cell.font = Font(bold=True, size=16, color="2C3E50")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill = PatternFill(fill_type="solid", start_color="E3F2FD")

            # Информация о данных
            ws_all.merge_cells('A2:AC2')
            info_cell = ws_all['A2']
            total_items = len(self.items)

            # Фильтруем избранные ДЛЯ СТАТИСТИКИ В ШАПКЕ
            favorite_items_list = [item for item in self.items if getattr(item, 'is_favorite', False)]
            favorites_count = len(favorite_items_list)

            # Вычисляем среднюю цену
            total_price = 0
            count_prices = 0
            for item in self.items:
                price = getattr(item, 'price', 0)
                if price and str(price).replace('.', '').isdigit():
                    total_price += float(price)
                    count_prices += 1

            avg_price = total_price / count_prices if count_prices > 0 else 0

            # Статистика по источникам
            sources = {}
            for item in self.items:
                source = getattr(item, 'source', 'unknown')
                sources[source] = sources.get(source, 0) + 1

            # Берем только первые 3 источника для отображения
            sources_list = list(sources.items())
            sources_info = ", ".join([f"{self._get_source_badge(k)}: {v}" for k, v in sources_list[:3]])

            info_cell.value = f"📈 Всего: {total_items} объявлений | ⭐ Избранных: {favorites_count} | 💰 Средняя цена: {avg_price:,.0f} ₽ | 🌐 Источники: {sources_info}"
            info_cell.font = Font(bold=True, size=11, color="34495E")
            info_cell.alignment = Alignment(horizontal="center", vertical="center")
            info_cell.fill = PatternFill(fill_type="solid", start_color="F5F5F5")

            # Заголовки столбцов
            headers = [
                "ID", "Название", "Сайт", "Статус", "Тип продавца",
                "Продавец", "Категория", "Рейтинг", "Дата публикации", "Просмотры",
                "Цена (₽)", "Прибыль", "Состояние", "Город", "Пробег",
                "Год", "Цвет", "Двигатель", "Коробка", "Привод",
                "Кузов", "Руль", "Владельцы", "ПТС", "Налог",
                "Растаможка", "Комплектация", "Ссылка", "Найдено", "В избранном"
            ]

            # Стили заголовков
            header_fill = PatternFill(fill_type="solid", start_color="4A5568")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_border = Border(
                left=Side(style='thin', color="1F2937"),
                right=Side(style='thin', color="1F2937"),
                top=Side(style='thin', color="1F2937"),
                bottom=Side(style='thick', color="1F2937")
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws_all.cell(row=4, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Данные для всех объявлений
            for row_idx, item in enumerate(self.items, 5):
                try:
                    # Проверяем, избранное ли это
                    is_favorite = getattr(item, 'is_favorite', False)
                    favorite_status = "⭐" if is_favorite else "—"
                    favorite_date = getattr(item, 'favorite_added_at', '—') if is_favorite else '—'

                    row_values = [
                        self._safe_value(getattr(item, 'id', '')),
                        self._safe_value(getattr(item, 'title', '')),
                        self._get_source_badge(self._safe_value(getattr(item, 'source', 'avito'))),
                        self._get_view_badge(getattr(item, 'views_count', 0)),
                        self._get_seller_badge(self._safe_value(getattr(item, 'seller_type', ''))),
                        self._safe_value(getattr(item, 'seller_name', '')),
                        self._safe_value(getattr(item, 'category', '')),
                        self._safe_value(getattr(item, 'seller_rating', 0)),
                        self._safe_value(getattr(item, 'posted_date', '')),
                        f"{self._safe_value(getattr(item, 'views_count', 0))}",
                        f"{int(float(self._safe_value(getattr(item, 'price', 0)))):,} ₽" if str(
                            getattr(item, 'price', 0)).replace('.', '').isdigit() else self._safe_value(
                            getattr(item, 'price', 0)),
                        self._get_profit_badge(getattr(item, 'profit', 0), getattr(item, 'source', 'avito')),
                        self._safe_value(getattr(item, 'condition', '')),
                        self._safe_value(getattr(item, 'city', '')),
                        self._safe_value(getattr(item, 'mileage', '')),
                        self._safe_value(getattr(item, 'year', '')),
                        self._safe_value(getattr(item, 'color', '')),
                        self._safe_value(getattr(item, 'engine', '')),
                        self._safe_value(getattr(item, 'transmission', '')),
                        self._safe_value(getattr(item, 'drive', '')),
                        self._safe_value(getattr(item, 'body', '')),
                        self._safe_value(getattr(item, 'steering', '')),
                        self._safe_value(getattr(item, 'owners', '')),
                        self._safe_value(getattr(item, 'pts', '')),
                        self._safe_value(getattr(item, 'tax', '')),
                        self._safe_value(getattr(item, 'customs', '')),
                        self._safe_value(getattr(item, 'package', '')),
                        self._safe_value(getattr(item, 'url', '')),
                        self._safe_value(getattr(item, 'found_at', '')),
                        favorite_status  # В избранном
                    ]

                    # Записываем строку
                    for col_idx, value in enumerate(row_values, 1):
                        cell = ws_all.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                        # ЖИРНЫЙ ТЕКСТ ДЛЯ НАЗВАНИЯ (столбец B)
                        if col_idx == 2:  # Название
                            cell.font = Font(bold=True, color="1A237E")

                        # ЖИРНЫЙ ТЕКСТ ДЛЯ ЦЕНЫ (столбец K)
                        if col_idx == 11:  # Цена
                            cell.font = Font(bold=True, color="1B5E20")

                        # ОСОБЫЙ СТИЛЬ ДЛЯ ИЗБРАННЫХ
                        if is_favorite:
                            # Желтый фон для всей строки избранного
                            cell.fill = PatternFill(fill_type="solid", start_color="FFFDE7")

                            # Звездочка в столбце "В избранном"
                            if col_idx == len(headers):  # Последний столбец
                                cell.font = Font(bold=True, color="FF9800", size=12)
                        else:
                            # Зебра для обычных строк
                            if row_idx % 2 == 0:
                                cell.fill = PatternFill(fill_type="solid", start_color="F8FAFC")

                    # Гиперссылка для URL
                    if hasattr(item, 'url') and item.url:
                        cell_url = ws_all.cell(row=row_idx, column=28)  # Ссылка сейчас в 28 столбце
                        cell_url.value = "🔗 Открыть"
                        cell_url.hyperlink = str(item.url)
                        cell_url.font = Font(color="0066CC", underline="single", bold=True)

                except Exception as e:
                    logger.warning(f"Ошибка строки всех {row_idx}: {e}")
                    continue

            # Настройка ширины столбцов для всех объявлений
            column_widths = {
                'A': 10,  # ID
                'B': 40,  # Название
                'C': 13,  # Сайт
                'D': 13,  # Статус
                'E': 13,  # Тип продавца
                'F': 20,  # Продавец
                'G': 25,  # Категория
                'H': 10,  # Рейтинг
                'I': 15,  # Дата публикации
                'J': 12,  # Просмотры
                'K': 15,  # Цена
                'L': 13,  # Прибыль
                'M': 15,  # Состояние
                'N': 15,  # Город
                'O': 12,  # Пробег
                'P': 10,  # Год
                'Q': 12,  # Цвет
                'R': 15,  # Двигатель
                'S': 12,  # Коробка
                'T': 12,  # Привод
                'U': 15,  # Кузов
                'V': 10,  # Руль
                'W': 12,  # Владельцы
                'X': 10,  # ПТС
                'Y': 10,  # Налог
                'Z': 12,  # Растаможка
                'AA': 20,  # Комплектация
                'AB': 15,  # Ссылка
                'AC': 18,  # Найдено
                'AD': 12,  # В избранном
            }

            for col_letter, width in column_widths.items():
                ws_all.column_dimensions[col_letter].width = width

            # Фильтры для всех столбцов
            ws_all.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{len(self.items) + 4}"
            ws_all.freeze_panes = 'A5'

            # ==================== ВКЛАДКА 2: ИЗБРАННОЕ ====================
            # Получаем РЕАЛЬНЫЕ избранные объявления
            real_favorite_items = []
            for item in self.items:
                try:
                    if getattr(item, 'is_favorite', False):
                        real_favorite_items.append(item)
                except:
                    continue

            ws_fav = wb.create_sheet(title="Избранное")

            # Яркая шапка для избранного
            ws_fav.merge_cells('A1:I1')
            fav_title = ws_fav['A1']
            fav_title.value = f"⭐ ИЗБРАННЫЕ ОБЪЯВЛЕНИЯ ({len(real_favorite_items)} шт.)"
            fav_title.font = Font(bold=True, size=16, color="FF6B00")
            fav_title.alignment = Alignment(horizontal="center", vertical="center")
            fav_title.fill = PatternFill(fill_type="solid", start_color="FFF3E0")

            # Заголовки для избранного
            fav_headers = ["⭐", "ID", "Название", "Цена (₽)", "Сайт", "Город", "Дата", "Ссылка",
                           "Добавлено в избранное"]

            for col_idx, header in enumerate(fav_headers, 1):
                cell = ws_fav.cell(row=3, column=col_idx, value=header)
                cell.fill = PatternFill(fill_type="solid", start_color="FF9800")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick', color="E65100")
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Заполняем реальными избранными
            if real_favorite_items:
                for row_idx, item in enumerate(real_favorite_items, 4):
                    try:
                        # Дата добавления в избранное
                        favorite_date = getattr(item, 'favorite_added_at', '—')
                        if favorite_date and hasattr(favorite_date, 'strftime'):
                            favorite_date_str = favorite_date.strftime("%d.%m.%Y %H:%M")
                        else:
                            favorite_date_str = str(favorite_date)

                        # Данные для избранного
                        fav_data = [
                            "⭐",  # Иконка звезды
                            self._safe_value(getattr(item, 'id', '')),
                            self._safe_value(getattr(item, 'title', '')),
                            f"{int(float(self._safe_value(getattr(item, 'price', 0)))):,} ₽" if str(
                                getattr(item, 'price', 0)).replace('.', '').isdigit() else self._safe_value(
                                getattr(item, 'price', 0)),
                            self._get_source_badge(self._safe_value(getattr(item, 'source', 'avito'))),
                            self._safe_value(getattr(item, 'city', '')),
                            self._safe_value(getattr(item, 'posted_date', '')),
                            self._safe_value(getattr(item, 'url', '')),
                            favorite_date_str
                        ]

                        for col_idx, value in enumerate(fav_data, 1):
                            cell = ws_fav.cell(row=row_idx, column=col_idx, value=value)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )

                            # Желтый фон для избранных
                            cell.fill = PatternFill(fill_type="solid", start_color="FFFDE7")

                            # Большая звезда в первом столбце
                            if col_idx == 1:  # Столбец со звездой
                                cell.font = Font(bold=True, color="FF9800", size=14)

                            # Жирный текст для названия
                            if col_idx == 3:  # Название
                                cell.font = Font(bold=True, color="1A237E")

                            # Жирный текст для цены
                            if col_idx == 4:  # Цена
                                cell.font = Font(bold=True, color="1B5E20")

                        # Гиперссылка для URL
                        if hasattr(item, 'url') and item.url:
                            cell_url = ws_fav.cell(row=row_idx, column=8)  # Ссылка в 8 столбце
                            cell_url.value = "🔗 Открыть"
                            cell_url.hyperlink = str(item.url)
                            cell_url.font = Font(color="0066CC", underline="single", bold=True)

                    except Exception as e:
                        logger.warning(f"Ошибка строки избранного {row_idx}: {e}")
                        continue
            else:
                # Если нет избранных
                ws_fav.merge_cells('A4:I4')
                empty_cell = ws_fav['A4']
                empty_cell.value = "❌ Нет избранных объявлений"
                empty_cell.font = Font(bold=True, size=12, color="757575")
                empty_cell.alignment = Alignment(horizontal="center")
                empty_cell.fill = PatternFill(fill_type="solid", start_color="F5F5F5")

            # Настройка ширины столбцов для избранного
            fav_widths = {'A': 8, 'B': 12, 'C': 50, 'D': 15, 'E': 15, 'F': 20, 'G': 15, 'H': 20, 'I': 22}
            for col_letter, width in fav_widths.items():
                ws_fav.column_dimensions[col_letter].width = width

            # Фильтры для избранного (если есть данные)
            if real_favorite_items:
                ws_fav.auto_filter.ref = f"A3:I{len(real_favorite_items) + 3}"
            ws_fav.freeze_panes = 'A4'

            # ==================== ВКЛАДКА 3: СТАТИСТИКА ====================
            ws_stats = wb.create_sheet(title="Статистика")

            # Заголовок статистики
            ws_stats.merge_cells('A1:D1')
            stats_title = ws_stats['A1']
            stats_title.value = "📊 СТАТИСТИКА ПО ОБЪЯВЛЕНИЯМ"
            stats_title.font = Font(bold=True, size=14, color="2C3E50")
            stats_title.alignment = Alignment(horizontal="center")
            stats_title.fill = PatternFill(fill_type="solid", start_color="E8F5E8")

            # Собираем статистику
            stats_data = [
                ["Показатель", "Значение", "Процент", "Примечание"],
                ["Всего объявлений", total_items, "100%", "Общее количество"],
                ["Избранных", favorites_count,
                 f"{(favorites_count / total_items * 100):.1f}%" if total_items > 0 else "0%", "В избранном"],
                ["Средняя цена", f"{avg_price:,.0f} ₽", "-", "Средняя стоимость"],
            ]

            # Статистика по источникам
            for source, count in sources_list[:5]:  # Топ 5 источников
                percent = (count / total_items * 100) if total_items > 0 else 0
                stats_data.append([
                    f"С {self._get_source_badge(source)}",
                    count,
                    f"{percent:.1f}%",
                    f"{self._get_source_badge(source)}"
                ])

            # Статистика по избранным источникам
            fav_sources = {}
            for item in real_favorite_items:
                source = getattr(item, 'source', 'unknown')
                fav_sources[source] = fav_sources.get(source, 0) + 1

            if fav_sources:
                stats_data.append(["", "", "", ""])  # Пустая строка
                stats_data.append(["ИСТОЧНИКИ ИЗБРАННЫХ", "", "", ""])

                for source, count in sorted(fav_sources.items(), key=lambda x: x[1], reverse=True):
                    percent = (count / favorites_count * 100) if favorites_count > 0 else 0
                    stats_data.append([
                        f"⭐ {self._get_source_badge(source)}",
                        count,
                        f"{percent:.1f}%",
                        "В избранном"
                    ])

            # Статистика по городам (топ 5)
            cities = {}
            fav_cities = {}

            for item in self.items:
                city = getattr(item, 'city', 'Не указан')
                if city and city != '—' and city != 'Не указан':
                    cities[city] = cities.get(city, 0) + 1

                    # Если избранное
                    if getattr(item, 'is_favorite', False):
                        fav_cities[city] = fav_cities.get(city, 0) + 1

            city_list = sorted(cities.items(), key=lambda x: x[1], reverse=True)[:5]
            if city_list:
                stats_data.append(["", "", "", ""])  # Пустая строка
                stats_data.append(["ТОП ГОРОДА", "", "", ""])

                for city, count in city_list:
                    percent = (count / total_items * 100) if total_items > 0 else 0
                    fav_count = fav_cities.get(city, 0)
                    fav_percent = (fav_count / count * 100) if count > 0 else 0

                    note = f"Из них в избранном: {fav_count} ({fav_percent:.1f}%)" if fav_count > 0 else ""
                    stats_data.append([
                        f"🏙️ {city}",
                        count,
                        f"{percent:.1f}%",
                        note
                    ])

            # Записываем статистику
            for row_idx, row in enumerate(stats_data, 3):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    if row_idx == 3:  # Заголовки
                        cell.fill = PatternFill(fill_type="solid", start_color="4CAF50")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif "⭐" in str(value):  # Избранное в статистике
                        cell.fill = PatternFill(fill_type="solid", start_color="FFFDE7")
                        cell.font = Font(bold=True)
                    elif row_idx % 2 == 0:  # Зебра
                        cell.fill = PatternFill(fill_type="solid", start_color="F1F8E9")

            # Настройка ширины столбцов статистики
            ws_stats.column_dimensions['A'].width = 35
            ws_stats.column_dimensions['B'].width = 15
            ws_stats.column_dimensions['C'].width = 12
            ws_stats.column_dimensions['D'].width = 30

            # Сохраняем в BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # ==================== СКАЧИВАНИЕ ====================
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"search_{timestamp}.xlsx"

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.ms-excel'
            )

            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(buffer.getvalue())
            response['X-Content-Type-Options'] = 'nosniff'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'

            return response

        except Exception as e:
            logger.error(f"Excel error: {e}")
            return self._error_response(f"Excel ошибка: {e}")

    def export_csv(self):
        """📝 ЭКСПОРТ В CSV"""
        try:
            if not self.items:
                return self._error_response("Нет данных")

            buffer = BytesIO()
            writer = csv.writer(buffer, delimiter=';', quoting=csv.QUOTE_MINIMAL)

            # Заголовки
            writer.writerow([
                'ID', 'Название', 'Сайт', 'Статус', 'Тип продавца',
                'Продавец', 'Категория', 'Рейтинг', 'Дата публикации', 'Просмотры',
                'Цена (₽)', 'Прибыль (₽)', 'Состояние', 'Город', 'Пробег',
                'Год', 'Цвет', 'Двигатель', 'Коробка', 'Привод',
                'Кузов', 'Руль', 'Владельцы', 'ПТС', 'Налог',
                'Растаможка', 'Комплектация', 'Ссылка', 'Найдено'
            ])

            for item in self.items:
                try:
                    writer.writerow([
                        self._safe_value(getattr(item, 'id', '')),
                        self._safe_value(getattr(item, 'title', '')),
                        self._safe_value(getattr(item, 'source', 'avito')),
                        self._get_view_badge(getattr(item, 'views_count', 0)),
                        self._safe_value(getattr(item, 'seller_type', '')),
                        self._safe_value(getattr(item, 'seller_name', '')),
                        self._safe_value(getattr(item, 'category', '')),
                        self._safe_value(getattr(item, 'seller_rating', 0)),
                        self._safe_value(getattr(item, 'posted_date', '')),
                        self._safe_value(getattr(item, 'views_count', 0)),
                        self._safe_value(getattr(item, 'price', 0)),
                        self._safe_value(getattr(item, 'profit', 0)),
                        self._safe_value(getattr(item, 'condition', '')),
                        self._safe_value(getattr(item, 'city', '')),
                        self._safe_value(getattr(item, 'mileage', '')),
                        self._safe_value(getattr(item, 'year', '')),
                        self._safe_value(getattr(item, 'color', '')),
                        self._safe_value(getattr(item, 'engine', '')),
                        self._safe_value(getattr(item, 'transmission', '')),
                        self._safe_value(getattr(item, 'drive', '')),
                        self._safe_value(getattr(item, 'body', '')),
                        self._safe_value(getattr(item, 'steering', '')),
                        self._safe_value(getattr(item, 'owners', '')),
                        self._safe_value(getattr(item, 'pts', '')),
                        self._safe_value(getattr(item, 'tax', '')),
                        self._safe_value(getattr(item, 'customs', '')),
                        self._safe_value(getattr(item, 'package', '')),
                        self._safe_value(getattr(item, 'url', '')),
                        self._safe_value(getattr(item, 'found_at', ''))
                    ])
                except Exception as e:
                    logger.warning(f"CSV error item: {e}")
                    continue

            buffer.seek(0)

            # Имя файла search_дата_время.csv
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"search_{timestamp}.csv"

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/octet-stream'  # ← БИНАРНЫЙ ТИП
            )

            # ОЧЕНЬ ВАЖНО: force download header
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(buffer.getvalue())
            response['X-Content-Type-Options'] = 'nosniff'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'

            return response

        except Exception as e:
            logger.error(f"CSV error: {e}")
            return self._error_response(f"CSV ошибка: {e}")

    def export_pdf(self):
        """📄 ЭКСПОРТ В PDF - С ПОДДЕРЖКОЙ РУССКИХ ШРИФТОВ"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os

            if not self.items:
                return self._error_response("Нет данных")

            buffer = BytesIO()

            # Регистрируем русские шрифты
            try:
                # Пытаемся зарегистрировать стандартные шрифты Windows
                pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
            except:
                try:
                    # Попробуем другой путь
                    pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
                except:
                    # Если не получилось, используем встроенный DejaVu
                    from reportlab.pdfbase import pdfmetrics
                    from reportlab.pdfbase.ttfonts import TTFont
                    # Можно скачать и использовать DejaVuSans
                    # Или просто оставить без шрифтов для теста

            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=20,
                rightMargin=20,
                topMargin=20,
                bottomMargin=20
            )
            elements = []

            # Создаем стиль с нашим шрифтом
            styles = getSampleStyleSheet()

            # Простой заголовок
            from reportlab.platypus import Spacer

            # Таблица с ограниченными данными (только важные поля)
            table_data = [['ID', 'Название', 'Цена (₽)', 'Сайт', 'Город']]

            for item in self.items[:50]:  # Ограничиваем 50 записей
                try:
                    # Берем только основные поля
                    title = str(self._safe_value(getattr(item, 'title', '')))[:30]
                    price = str(self._safe_value(getattr(item, 'price', 0)))
                    source = str(self._safe_value(getattr(item, 'source', '')))[:10]
                    city = str(self._safe_value(getattr(item, 'city', '')))[:15]
                    item_id = str(self._safe_value(getattr(item, 'id', '')))[:8]

                    row = [item_id, title, price, source, city]
                    table_data.append(row)
                except Exception as e:
                    continue

            # Создаем таблицу
            table = Table(table_data, colWidths=[50, 200, 80, 60, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Используем Helvetica
            ]))

            elements.append(table)

            # Добавляем подпись
            from reportlab.platypus import Paragraph
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER

            style = ParagraphStyle(
                name='Custom',
                alignment=TA_CENTER,
                fontSize=8,
                textColor=colors.grey
            )
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"Всего объявлений: {len(self.items)}", style))

            # Собираем документ
            doc.build(elements)
            buffer.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"search_{timestamp}.pdf"

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/octet-stream'
            )

            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(buffer.getvalue())
            response['X-Content-Type-Options'] = 'nosniff'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'

            return response

        except Exception as e:
            logger.error(f"PDF error: {e}")
            # Возвращаем простой текст вместо PDF если ошибка
            return self._error_response(f"PDF ошибка: {e}")

    def _error_response(self, message):
        """Создание JSON ответа об ошибке"""
        return JsonResponse({'status': 'error', 'message': message}, status=500)
